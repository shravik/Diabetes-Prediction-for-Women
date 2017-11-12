from flask import Flask, render_template, request

import pandas as pd
import xlrd
import sklearn
import openpyxl
import os
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt
import numpy as np
import xlwt
from scipy import sparse
from xlutils.copy import copy


app = Flask(__name__)

@app.route('/', methods=['GET'])
def index():
    if(request.method == 'GET'):
        return render_template('user_input.html')


@app.route('/getdata', methods=['GET', 'POST'])
def check():
    print("I came here")

    preg = request.form.get('preganency')
    bloodP =request.form.get('BloodPressure')
    skinT = request.form.get('skinthickness')
    insulin = request.form.get('insulin')
    bmi = request.form.get('bmi')
    age = request.form.get('age')

    print("Preg = ", preg, " Blood Pressure= ", bloodP)
    print("==============================")

    appendDataToExcel(preg,bloodP,skinT, insulin, bmi, age)
    print("=========Text Excel Saved=======")

    var = precitMe()

    msg = ''
    if var < 150 and var >= 90:
        msg = "Your sugar level is normal! Keep up the good work"
    elif var >= 160 and var <= 240:
        msg = "Your blood sugar is too high. Work on lowering down your sugar intake"
    elif var >= 240 and var <= 300:
        msg = "Your sugar level is out of control"
    else:
        msg = "Your life is in danger"


    return render_template('display_data.html', msg = msg)

def appendDataToExcel(preg,bloodP,skinT, insulin, bmi, age):
    print("Entered Append Function")

    file = 'C:\\Users\\salon\\Desktop\\Final\\diabetes_data.xlsx'
    new_row = [preg,bloodP,skinT, insulin,bmi, age]
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.get_sheet_by_name('diabetes')
    row = ws.max_row +1
    for col, entry in enumerate(new_row, start=1):
        ws.cell(row=row, column=col, value=entry)
    wb.save(file)
    print("Leaving Append Function")



def precitMe():
    print('Into the function predictMe.... Started Function')
    value = predictDiabetes()
    var = int(value[0])
    wb = load_workbook(filename='C:\\Users\\salon\\Desktop\\Final\\diabetes_data.xlsx')
    ws = wb.worksheets[0]
    row = ws.max_row
    ws.cell(row=row, column=7).value = var
    print("====================================================")
    print(var)

    return var



@app.route('/foodItem', methods=['GET', 'POST'])
def foodItem():
    print("entered food Item")
    newValue = request.form['foodValue']
    print(newValue)
    print("======0eiqw8iaosjmk")
    database = xlrd.open_workbook('C:\\Users\\salon\\Desktop\\Final\\Foods.xlsx')
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.get_sheet_by_name('diabetes')
    sugar = ''

    print("11111")
    for i in range(1, ws.max_row):

        if ws.cell(row=i, column=1).value == newValue:
            print("entered this thing")
            print(sugar)
            sugar = ws.cell(row=i, column=3).value
    sugar = 2500-sugar
    print("22222222222")
    msg1= 'You need to have ' + sugar + ' for the day'
    msg = ''
    if sugar < 270 and sugar >= 135:
        msg = "No sugar for you! You are done for today"
    elif sugar >= 135 and sugar <= 95:
        msg = "You can have some Tofu, beans, oats"
    elif sugar >= 95 and sugar <= 50:
        msg = "You can have fruits"
    else:
        msg = "You can eat anything upto 2500 mg of sugar! "

    print("left food item")

    return render_template('final.html', msg = msg, msg1 = msg1 )




def predictDiabetes():
    print("lol i came here")
    database = xlrd.open_workbook('C:\\Users\\salon\\Desktop\\Final\\diabetes_data.xlsx')

    col_names = ['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'Age',
                 'Glucose']
    pima = pd.read_excel(database, names=col_names, engine='xlrd')
    per = 0.55
    plot = pd.DataFrame(col_names)
    train = pima[:((int)(len(plot) * per))]
    xtrain = train[
        ['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'Age']]
    ytrain = train[['Glucose']]

    test = pima[((int)(len(plot) * per)):]
    xtest = test[['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'Age']]
    ytest = test[['Glucose']]
    dataframeHeading = plot.head()

    ols = sklearn.linear_model.LinearRegression()
    model = ols.fit(xtrain, ytrain)

    predict = model.predict(xtest)
    print("==Value of model predict===")

    length = len(predict) -1
    print(predict[length])
    print("This is the value?")
    print("Ends Here")

    return predict[length]

if '__main__' == __name__:
    app.run()